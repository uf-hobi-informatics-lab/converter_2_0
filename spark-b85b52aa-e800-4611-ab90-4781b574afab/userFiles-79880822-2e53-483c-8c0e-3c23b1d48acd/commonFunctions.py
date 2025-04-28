
###################################################################################################################################
# This class will host all the common functions used in the mappings by diffrenet mapping scripts
###################################################################################################################################



from itertools import chain
from Crypto.Hash import SHA
from Crypto.Cipher import XOR
from base64 import b64decode, b64encode
from pyspark.sql.functions import udf, col, count, desc, upper, round,  lit, length, max
from pyspark.sql import SparkSession 
from pyspark import SparkConf, SparkContext
import pyspark
import time
import subprocess 
from pyspark.sql.types import StructType, StructField, StringType, NullType, DateType, FloatType, TimestampType, DecimalType, DoubleType
import sys
from ovid_secrets import * 
from settings import *
from pcornet_cdm import PcornetCDM
from pyspark.conf import SparkConf
import pandas as pd
from urllib import parse
from datetime import datetime
import pytz
from dateutil import parser as dp
import pyspark.sql.functions as F
import os
import importlib
import openpyxl
from PIL import Image
import re
import snowflake.connector





class CommonFuncitons:


    def __init__(self, partner):




        self.partner = partner # specifying the partner / used for the hashing
        self.get_time_from_datetime_udf = udf(self.get_time_from_datetime, StringType())
        self.get_date_from_datetime_udf = udf(self.get_date_from_datetime, StringType())
        self.format_date_udf = udf(self.format_date, StringType())
        self.format_time_udf = udf(self.format_time, StringType())
        self.encrypt_id_udf = udf(self.encrypt_id, StringType())
        self.get_current_time_udf = udf(self.get_current_time, StringType())
        self.get_date_from_date_str_with_default_value_udf = udf(self.get_date_from_date_str_with_default_value, StringType())
        self.get_datetime_from_date_str_with_default_value_udf = udf(self.get_datetime_from_date_str_with_default_value, StringType())
        self.copy_if_float_udf= udf(self.copy_if_float, StringType())
        self.return_if_float_udf = udf(self.return_if_float, StringType())
        






    


###################################################################################################################################
# This function will return the xor hashed value based on the partner name and input value/id
###################################################################################################################################

   
   

    # @classmethod
    def encrypt_id(cls, id):

            if id is None:

                return None
            else: 


                partner_encryption_value_dict = {

                        'AVH':'FLH',
                        'BND':'BND',
                        'CHP':'CHP',
                        'CHT':'CHT',
                        'EMY':'EMY',
                        'FLM':'FLM',
                        'NCH':'NCH',
                        'ORH':'ORH',
                        'ORL':'ORL',
                        'TMH':'TMH',
                        'TMA':'TMA',
                        'TMC':'TMC',
                        'UAB':'UAB',
                        'UFH':'UFH',
                        'UMI':'UMI',
                        'USF':'USF',
                        'TGH':'TGH',
                        'AMS':'AMS',
                        'UCI':'UCI',
                        'omop_partner_1'     :'omop_partner_1',
                        'omop_partner_plus'  :'omop_partner_plus',
                        'pcornert_partner_1' :'pcornert_partner_1',

                }

                partner_encryption_value = partner_encryption_value_dict.get(cls.partner.upper(), None)

                HASHING_KEY = SHA.new(SEED.encode('utf-8')).digest()

                cipher = XOR.new(HASHING_KEY)

                return b64encode(cipher.encrypt('{}{}'.format(partner_encryption_value, id))).decode()


###################################################################################################################################
# This function will return the current datetime
###################################################################################################################################

   

    @classmethod
    def get_current_time(cls):

       # Set the timezone to Eastern Time (New York)
        eastern_timezone = pytz.timezone("US/Eastern")

        # Get the current time in the Eastern Time Zone
        current_time = datetime.now(eastern_timezone)

        # Define the desired time format
        time_format = "%Y-%m-%d %H:%M:%S"

        # Format and print the current time
        formatted_time = current_time.strftime(time_format)

        return formatted_time


        
    
###################################################################################################################################
# This function will output a pyspark dataframe and combine all the sub files into one file and delete the temprary files
###################################################################################################################################

   
    @classmethod
    def write_pyspark_output_file(cls,payspark_df, output_file_name, output_data_folder_path ):

        work_directory = output_data_folder_path+"tmp_"+output_file_name

        payspark_df = payspark_df.select([
            F.lit(None).cast('string').alias(i.name)
            if isinstance(i.dataType, NullType)
            else i.name
            for i in payspark_df.schema
            ])


        payspark_df.write.format("csv").option("header", True).option("nullValue",None).option("delimiter", "\t").option("quote", "").mode("overwrite").save(work_directory)

        output_file_path = os.path.join(output_data_folder_path, output_file_name).replace('temp_','')
        
        # Update the awk command to use the correct Python variable for the output path
        command = f"awk 'FNR==1 && NR!=1{{next}} 1' {work_directory}/part-*.csv > {output_file_path}"
        subprocess.run(command, shell=True, check=True)

        command = ["rm","-r", work_directory]
        # Execute the command
        subprocess.run(" ".join(command), shell=True)


###################################################################################################################################
# This function will 
###################################################################################################################################

           
    def copy_data(cls,input_file_name,output_file_name,data_folder_path,output_data_folder_path ):

        if not os.path.exists(output_data_folder_path):
            os.makedirs(output_data_folder_path)

        command = ["cp", "-p", os.path.join(data_folder_path, input_file_name), os.path.join(output_data_folder_path,output_file_name)]
                # Execute the command
        subprocess.run(" ".join(command), shell=True)
###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   
    @classmethod
    def deduplicate(cls,partner,file_name,table_name, file_path,folder_name):

        spark = cls.get_spark_session('generate mapping report')

        deduplicated_files_path = file_path.replace('formatter', 'deduplicator')
        duplicates_file_path = os.path.join(deduplicated_files_path, 'duplicate_values')
        duplicates_file_name  = file_name.lower()+'_duplicates.csv'
        deduplicated_file_name  = 'deduplicated_'+file_name.lower()+'.csv'

        # Create the deduplicated files directory if it doesn't exist
        if not os.path.exists(deduplicated_files_path):
            os.makedirs(deduplicated_files_path)

        if not os.path.exists(duplicates_file_path):
            os.makedirs(duplicates_file_path)

        input_file = f"{file_path}formatted_{file_name.lower()}.csv"

        df = cls.spark_read(input_file, spark)
        
        row_count = df.count()



        if row_count > AWK_PASSBY_TRESHHOLD :

            spark = cls.get_spark_session('generate mapping report')

            try: 

                if file_name == 'ENROLLMENT':
                                
                    columns_to_consider = [df.columns[0], df.columns[1], df.columns[4]]

                            
                elif file_name == 'DEATH_CAUSE':
                                
                    columns_to_consider = [df.columns[0], df.columns[1],df.columns[2],df.columns[3], df.columns[4]]

                            
                elif file_name == 'PAT_RELATIONSHIP':
                                
                    columns_to_consider = [df.columns[0], df.columns[1],df.columns[2]]

                else:
                    columns_to_consider = [df.columns[0]]



                deduplicated_df = df.dropDuplicates(columns_to_consider)

                duplicates_df = df.groupBy(*columns_to_consider).agg(count("*").alias("count")).filter("count > 1")

                cls.write_pyspark_output_file(deduplicated_df,deduplicated_file_name, deduplicated_files_path )
                cls.write_pyspark_output_file(duplicates_df,duplicates_file_name, duplicates_file_path )

                spark.stop()

            except Exception as e:

                spark.stop()
                
                cls.print_with_style(str(e), 'danger red')



        else:

            # Set up the deduplicated files path

            spark.stop()

            deduplicated_output_file = f"{deduplicated_files_path}deduplicated_{file_name.lower()}.csv"
            duplicates_output_file = f"{duplicates_file_path}/{file_name.lower()}_duplicates.csv"



            # Construct the gawk script based on the file_name
            if file_name == 'ENROLLMENT':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $5]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $5] > 1 {{ print > "{duplicates_output_file}" }}
                '''
            
            elif file_name == 'DEATH_CAUSE':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $3, $4, $5]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $3, $4, $5] > 1 {{ print > "{duplicates_output_file}" }}
                '''
 
            elif file_name == 'PAT_RELATIONSHIP':
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1, $2, $3]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1, $2, $3] > 1 {{ print > "{duplicates_output_file}" }}
                
                '''          
            else:
                gawk_script = f'''
                BEGIN {{ FS = OFS = "\t" }}
                NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
                !seen[$1]++ {{ print > "{deduplicated_output_file}" }}
                seen[$1] > 1 {{ print > "{duplicates_output_file}" }}
                '''

            # Construct the full shell command
            command = f"awk '{gawk_script}' {input_file}"

            subprocess.run(command, shell=True, check=True)




###################################################################################################################################
# This function will 
###################################################################################################################################

    @classmethod
    def dv_match(cls,partner_name,folder,dv_mapping_table_path,fixed_table_path,table_name,output_path,dv_mapper_output_path ):

            spark = cls.get_spark_session('DV MATCH')


            if table_name != 'PROVIDER':

            

                try:

                    fixed_table      = cls.spark_read(fixed_table_path, spark)
                    dv_mapping_table = cls.spark_read(dv_mapping_table_path, spark)


                    joind_df = fixed_table.join(dv_mapping_table, dv_mapping_table["ENCRYPTED_PATID"] == fixed_table['PATID'], how = 'inner')

                    joind_df.show()

                except Exception as e:
                    cls.print_with_style(str(e), 'danger red')

                finally:
                    spark.stop()
            else:

                spark.stop()



###################################################################################################################################
# This function will 
###################################################################################################################################

    @classmethod

    def apply_fix(cls,fix_type, fix_name, version, input_path, output_path,fixed_table_name,mapped_table_name,input_data_folder,partner_name, src1, src2):

        if fix_type == 'custom':

            partner_fix_path = f"/app/partners/{partner_name.lower()}/custom_fixes/{fix_name}/{version}.py"


        if fix_type == 'common':



            partner_fix_path = f"/app/common/common_fixes/{fix_name}/{version}.py"

        command = ["python", partner_fix_path, '-f', input_data_folder, '-t', mapped_table_name , '-p', partner_name,'-ftm',fixed_table_name, '-i',input_path,'-o',output_path, '-sr1', src1, '-sr2', src2]

        subprocess.run(" ".join(command), shell=True)

###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   

    @classmethod
    def get_mapping_gap(cls,partner,file_name,table_name, file_path,folder_name):

        

        complete_file_path = file_path+'/formatted_'+file_name.lower()+'.csv'
        partner_dictionaries_path = "partners."+partner+".dictionaries"
        partner_dictionaries = importlib.import_module(partner_dictionaries_path)
        
        spark = SparkSession.builder.master("spark://master:7077").appName("MAPPING GAP").getOrCreate()

        df = cls.spark_read(complete_file_path, spark)
        df.cache()
        row_count= df.count()

        mapping_dicionaries_list = []

        for key, value in partner_dictionaries.__dict__.items():
            

            if  ((file_name.lower() in key and 'death_cause' not in key) or  (file_name.lower() in key and file_name.lower()=='death_cause') ) and "_dict" in key:
                
                # get the field name


                prefix = '@' #logic added to only remove the first word from the dictinary name

                temp_key = prefix+key


                field_name = temp_key.replace(prefix+file_name.lower()+'_',"").replace('_dict',"").upper()

                

                # get the aggreated data from the formatter for the field 

                # try:

                aggregated_values = cls.get_aggregate_value_from_table(df,row_count,field_name)


                # get and output the aggreated data that has mapping 

                dictionary_dataframe = spark.createDataFrame(value.items(), ["dict_key", "dict_value"])
        

                if 'UNIT' in field_name:
                    joined_df = aggregated_values.join(dictionary_dataframe, dictionary_dataframe['dict_key']==aggregated_values[field_name], how ="left").orderBy(desc("count"))
                else:
                    joined_df = aggregated_values.join(dictionary_dataframe, upper(dictionary_dataframe['dict_key'])==upper(aggregated_values[field_name]), how ="left").orderBy(desc("count"))
            

                cls.diagnos_dictionary(dictionary_dataframe, file_path, key, file_name.lower())

                cls.diagnos_mapping(joined_df, file_path, key, field_name, dictionary_dataframe, file_name.lower())

                mapping_dicionaries_list.append(key)     

        spark.stop()





###################################################################################################################################
# This function will diagnos a given dictionary data frame
########################################################################################################

    @classmethod
    def diagnos_dictionary(cls,dictionary_dataframe, file_path, file_name, table_name):

        # getting values that mapped to null or some flavor of null like 'OT, 'NI', and "UN
        filtered_df = dictionary_dataframe.filter((col("dict_value").isin("NI", "OT", "UN")) | (col("dict_value").isNull()) | (col("dict_value") == ""))

        out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/mapping_to_flavor_of_null/"
        output_file_name = file_name+"_flavor_of_null.csv"


        cls.write_pyspark_output_file(filtered_df, output_file_name, out_file_path )
           


###################################################################################################################################
# This function will diagnos a given joind dictionary dataframe
########################################################################################################

    @classmethod
    def diagnos_mapping(cls,joined_df, file_path, file_name, field_name, dictionary_dataframe, table_name):


        try:
            # output all mapping
            all_values_df = joined_df.select(joined_df[field_name],joined_df['dict_value'],joined_df['count'],joined_df['percent'])
            all_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/all_mapping/"
            all_mapping_output_file_name = file_name+"_all_mapping.csv"

            cls.write_pyspark_output_file(all_values_df, all_mapping_output_file_name, all_mapping_out_file_path )


            
            # output missing mapping
            cdm_value_set = PcornetCDM.get_cdm_valueset_df(field_name)
            cdm_value_set = [x.strip() for x in cdm_value_set] # remove extra spaces from the values list

            missing_values_df = all_values_df.filter(col("dict_value").isNull() )
            
            
            missing_values_df_and_not_part_of_cdm = missing_values_df.filter(~col(field_name).isin(cdm_value_set))
            
            missing_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/missing_mapping/"
            missing_mapping_output_file_name = file_name+"_missing_mapping.csv"
            cls.write_pyspark_output_file(missing_values_df_and_not_part_of_cdm, missing_mapping_output_file_name, missing_mapping_out_file_path )

            # output suggested mapping

            
            to_be_added_field = missing_values_df_and_not_part_of_cdm.select(

                    missing_values_df_and_not_part_of_cdm[field_name].alias('dict_key'),
                    lit('').alias('dict_value'),


            )

            unioned_df= dictionary_dataframe.union(to_be_added_field)
            suggested_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/suggested_mapping/"

            out_file_name = suggested_mapping_out_file_path+"suggested_"+file_name+".py"

            result = os.makedirs(os.path.dirname(out_file_name), exist_ok=True)

            formatted_text = unioned_df.rdd \
                .map(lambda row: f'       {cls.process_key(row["dict_key"],row["dict_value"],file_name, 20)}  :  "{row["dict_value"]}",') 
            

            formatted_text= formatted_text.collect()
            

            # Join the formatted strings with newline characters
            formatted_text = "\n".join(formatted_text)

            # Write the formatted text to a text file
        
            first_line = "\n\n"+file_name+ " = {\n\n"
            last_line = "   \n\n}"

            with open(out_file_name, "w+") as file:
                file.write(first_line)
                file.write(formatted_text)
                file.write(last_line)


        except Exception as e:

            

            cls.print_with_style(str(e), 'danger red')
            
            
    
###################################################################################################################################
# 
########################################################################################################

    @classmethod
    def process_row(cls,row, file):


            print(row)


            key = cls.append_spaces(row.dict_key, 20),

            text = row.dict_value + '\n'

            file.write(text)


                
               

###################################################################################################################################
# 
########################################################################################################

    @classmethod
    def process_key(cls, key, value,file_name, size):

            key = str(key)

            if 'unit' not in file_name and 'UNIT' not in file_name:

                key = key.upper()

            key = '"'+key+'"'


            if value == "" or value == None:

                key = '# '+key

            while len(key) < size:

                key = key + ' '

            return key





###################################################################################################################################
# This function return and aggreated dataframe for a given table
###################################################################################################################################



    @classmethod
    def get_aggregate_value_from_table(cls, df, row_count, field_name):
        aggregated_df = df.groupBy(field_name).agg(count("*").alias("count"))

        if row_count != 0:
            percent_col = round((col("count") / row_count) * 100, 2).alias("percent")
        else:
            percent_col = lit(0).alias("percent")

        return aggregated_df.select(col(field_name), col("count"), percent_col)





###################################################################################################################################
# This function will
###################################################################################################################################

    @classmethod
    def db_upload(cls,db, db_server,schemas,file_name,table_name, file_path,folder_name,db_type):

        if db_type.upper() == 'PG':
            cls.db_upload_postgress(db, db_server,schemas[0],file_name,table_name, file_path,folder_name)

        if db_type.upper() == 'MSSQL':
            cls.db_upload_mssql(db, db_server,schemas[0],file_name,table_name, file_path,folder_name)

        if db_type.upper() == 'SF':
            cls.db_upload_snowflake(db, db_server,schemas[1],file_name,table_name, file_path,folder_name)





###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

    @classmethod
    def db_upload_snowflake(cls,db, db_server,schema,file_name,table_name, file_path,folder_name):
            
            
     
            SNOWFLAKE_SOURCE_NAME = "net.snowflake.spark.snowflake"
           
            try:

                spark = SparkSession.builder \
                    .master("spark://master:7077") \
                    .appName("Upload") \
                    .getOrCreate()            

                custom_schema = StructType(schema)

                file_df = spark.read.load(file_path+file_name+'*', format="csv", sep="\t", schema=custom_schema, header="true", quote='"')
          
                conn = snowflake.connector.connect(
                    user=SF_DB_USER,
                    password=SF_DB_PASS,
                    account=db_server,
                    database=db
                )

                cur = conn.cursor()
                cur.execute(f"CREATE SCHEMA IF NOT EXISTS  {db}.OVID_{folder_name}")  # Set schema explicitly
                cur.close()
                conn.close()


                sf_options = {
                    "sfURL": "https://"+db_server+".snowflakecomputing.com/",
                    "sfDatabase": db,
                    "sfSchema": f"OVID_{folder_name}",
                    "sfRole": SF_ROLE,  # Optional
                    "sfUser": SF_DB_USER,
                    "sfPassword": SF_DB_PASS  # Add your password here
                }
                

                
                # Write DataFrame to Snowflake
                file_df.repartition(10).write \
                    .format(SNOWFLAKE_SOURCE_NAME)\
                    .options(**sf_options) \
                    .option("dbtable", table_name.upper()) \
                    .option("createTableColumnTypes", schema)\
                    .option("useSchema", f"OVID_{folder_name}")\
                    .mode("overwrite") \
                    .save()
                
                
                spark.stop()
                    
      
            except Exception as e:

                spark.stop()
                

                cls.print_with_style(str(e), 'danger red')





###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

    @classmethod
    def db_upload_mssql(cls,db, db_server,schema,file_name,table_name, file_path,folder_name):


        try:

            spark = SparkSession.builder.master("spark://master:7077").appName("MSSQL Upload").getOrCreate()

            conf = SparkConf()
            conf.set("spark.driver.extraClassPath", "mssql-jdbc-driver.jar")
            file_df = spark.read.option("inferSchema", "false").load(file_path+file_name+'*',format="csv", sep="\t", inferSchema="false", header="true",  quote= '"')
            
            string_columns = [f'CAST(`{col}` AS STRING) AS `{col}`' for col in file_df.columns]
            result_df = file_df.selectExpr(*string_columns)

            
            jdbc_url = f"jdbc:sqlserver://{db_server};databaseName={db}"
            
            if schema != None:
                properties = {
                    "user": MSSQL_DB_USER,
                    "password": MSSQL_DB_PASS,
                    "driver": "com.microsoft.sqlserver.jdbc.SQLServerDriver",
                    "trustServerCertificate": "true",
                    "createTableColumnTypes": schema}
            else:
                properties = {
                    "user": MSSQL_DB_USER,
                    "password": MSSQL_DB_PASS,
                    "driver": "com.microsoft.sqlserver.jdbc.SQLServerDriver",
                    "trustServerCertificate": "true"}


            table_name = f"OVID_{folder_name}_{table_name}"
            
            result_df.write.jdbc(url=jdbc_url, table=table_name, mode="overwrite", properties=properties)

            spark.stop()
        except Exception as e:

            spark.stop()
            

            cls.print_with_style(str(e), 'danger red')




###################################################################################################################################
# This function will upload a pyspark df to postgress table in a specified server/dabase
###################################################################################################################################

    @classmethod
    def db_upload_postgress(cls,db, db_server,schema,file_name,table_name, file_path,folder_name):
                #Boot Cluster
        start = time.time()
        date_format_pattern = "yyyy-MM-dd"

        spark = SparkSession.builder.master("spark://master:7077").appName("PSQL Upload").getOrCreate()

        try:
            conf = SparkConf()
            conf.set("spark.driver.extraClassPath", "postgresql-42.7.4.jar")

            # Parse schema to get column names and types
            column_types = {}
            for line in schema.strip().split(",\n"):
                match = re.match(r"(\w+)\s+(\w+)", line.strip())
                if match:
                    column, dtype = match.groups()
                    column_types[column] = dtype

            # Load your file into a DataFrame
            file_df = spark.read.option("inferSchema", "false").csv(file_path + file_name + '*', sep="\t", header=True, quote='"')


            # Create casting expressions based on parsed schema
            casting_expressions = [
                f'CAST(`{col}` AS {"STRING" if dtype.startswith("VARCHAR") else dtype}) AS `{col}`'
                for col, dtype in column_types.items()
            ]

            # Apply casting expressions
            result_df = file_df.selectExpr(*casting_expressions)
            
            if schema != None:
                properties = {
                    "user": PG_DB_USER,
                    "password": PG_DB_PASS,
                    "driver": "org.postgresql.Driver",
                   # "trustServerCertificate": "true",
                    "stringtype":"unspecified",
                    "createTableColumnTypes": schema}
            else:
                properties = {
                    "user": PG_DB_USER,
                    "password": PG_DB_PASS,
                    "driver": "org.postgresql.Driver",
                   # "trustServerCertificate": "true"
                   }


            table_name = f"OVID_{folder_name}_{table_name}"
            jdbc_url = f"jdbc:postgresql://{db_server}/{db}"
            result_df.write.jdbc(url=jdbc_url, table=table_name, mode="overwrite", properties=properties)

            spark.stop()
        except Exception as e:

             spark.stop()
            

             cls.print_with_style(str(e), 'danger red')



###################################################################################################################################
# This function will return a list of folder in a given path
###################################################################################################################################
    @classmethod
    def get_folders_list(cls, path):

       
        
        folders_names = os.listdir(path)

        folders_names_list = [file for file in folders_names if file]

        return folders_names_list


###################################################################################################################################
# This function clean strings from values not allowed in Excel
###################################################################################################################################
    @classmethod
    def remove_illegal_chars(cls, value):
        if isinstance(value, str):
            # Remove any illegal characters (control characters: 0x00–0x1F except tab, newline, carriage return)
            return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', value)
        return value
        
###################################################################################################################################
# This function generate a mapping report as a form of an excel file
###################################################################################################################################


    @classmethod
    def generate_mapping_report(cls,mapping_gap_output_folder_path, folder, partner):

        spark = cls.get_spark_session('generate mapping report')

        empty_columns = pd.DataFrame(columns=[""] * 2)
        # Path to the Excel file
        excel_file_path = mapping_gap_output_folder_path + partner+'_'+folder+"_mapping_report.xlsx"

        # Initialize a workbook object
        wb = openpyxl.Workbook()



        tables_list = cls.get_folders_list(mapping_gap_output_folder_path)

        for table in tables_list :

           


            if 'mapping_report' not in table:

                ws = wb.create_sheet(title=table)

                combined_pd = empty_columns


                table_all_mappings_path = mapping_gap_output_folder_path +"/"+table+"/all_mapping/"

                table_all_mapping_files_list = cls.get_folders_list(table_all_mappings_path)

                for mapping_file in table_all_mapping_files_list:

                    mapping_df =  cls.spark_read(table_all_mappings_path+mapping_file, spark)

                    mapping_df = mapping_df.limit(1000000)


                    pandas_df = mapping_df.toPandas()
                    combined_pd = pd.concat([combined_pd, empty_columns, pandas_df], axis=1)


                    for col_idx, col_name in enumerate(combined_pd.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)


                    for row_idx, row in enumerate(combined_pd.values, 2):  # Start from row 2 (after header row)
                        for col_idx, value in enumerate(row, 1):

                            safe_value = cls.remove_illegal_chars(value)

                            
                            cell = ws.cell(row=row_idx, column=col_idx, value=safe_value)

                            # Calculate the width of the cell based on the length of the value
                            if isinstance(safe_value, str):
                                width = len(safe_value)
                            else:
                                width = len(str(safe_value))

                            # Update the column width if necessary
                            if width > ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width:
                                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        if 'Sheet' in wb.sheetnames:
            # Remove the sheet
            wb.remove(wb['Sheet'])

        wb.save(excel_file_path)




        spark.stop()



###################################################################################################################################
# This function will return a pysark dataframe for a givin file path
###################################################################################################################################


    @classmethod
    def spark_read(cls,file_path, spark):

        return spark.read.option("inferSchema", "false").load(file_path,format="csv",   sep="\t", inferSchema="false", header="true",  quote= '"')



###################################################################################################################################
# This function will append any additional fields that are not in the Pcornet CDM to the mapper output
###################################################################################################################################

   

    @classmethod
    def append_additional_fields(cls,mapped_df, file_name, deduplicated_data_folder_path,join_field, spark ):


        # read the deduplicated file
        deduplicated_file = spark.read.load(deduplicated_data_folder_path+file_name,format="csv", sep="\t", inferSchema="true", header="true", quote= '"')
        
        if len(deduplicated_file.columns) + 3 > len(mapped_df.columns) :

            #get the headers from the deduplicated and the mapped files for comparison
            deduplicated_file_columns = deduplicated_file.columns
            mapped_df_columns = mapped_df.columns


            # Identify common headers
            common_headers = set(mapped_df_columns).intersection(deduplicated_file_columns)
            common_headers_without_the_join_field = [item for item in common_headers if item != join_field]

            # Drop common headers from the second DataFrame
            additional_fields_data = deduplicated_file.drop(*common_headers_without_the_join_field)
            
            #Renaming the join field to avoid conflect during the join
            join_field_temp_name = join_field+"_new"

            additional_fields_data = additional_fields_data.withColumnRenamed(join_field, join_field_temp_name)
            
            # Join DataFrames based on the "Name" column
            appended_payspark_data_frame = mapped_df.join(additional_fields_data, mapped_df["JOIN_FIELD"]==additional_fields_data[join_field_temp_name])
        
        #Dropping the temporary join fields

            columns_to_drop = ["JOIN_FIELD",join_field_temp_name ]
            appended_payspark_data_frame = appended_payspark_data_frame.drop(*columns_to_drop)

            return appended_payspark_data_frame
        
        else:

            return mapped_df.drop("JOIN_FIELD")
        


###################################################################################################################################
# This function will convert all known time formats to the '%H:%M'format
###################################################################################################################################

    @staticmethod
    def format_time( val):
        """ Convenience method to try all known formats """

        try:
            parsed_time = dp.parse(val)
        except Exception as exc:
           
            return ""
        return parsed_time.strftime('%H:%M')


###################################################################################################################################
# This will ensure the value is a float or it will return None
###################################################################################################################################

    @staticmethod
    def copy_if_float( val):
        """ Convenience method to try all known formats """
        if not val:
            return None
        try:
            return float(val)
        except:

            return None



###################################################################################################################################
# This function will convert all known date formats to the '%Y-%m-%d format
###################################################################################################################################

   
    @staticmethod
    def format_date( val):
        """ Convenience method to try all known formats """
        print(val)
        try:
            parsed_date = dp.parse(val)
        except:

            print(val)
           
            return None

        return parsed_date.strftime('%Y-%m-%d')




###################################################################################################################################
# This function will return the time portion from a datetime value
###################################################################################################################################

   
    @staticmethod
    def get_time_from_datetime(  dateTime):
        
            try:
                
                return str(dateTime.time().strftime('%H:%M:%S'))[0:5]
            except:
                return None

###################################################################################################################################
# This function will return the date portion from a datetime value
###################################################################################################################################

   
    @staticmethod
    def get_date_from_datetime(  dateTime):
        
        try:
            
            return str(dateTime.strftime("%Y-%m-%d"))
        except:
            return None

        
  
###################################################################################################################################
# This function will validate if a giving partenr name is valid or not
###################################################################################################################################

   

    @classmethod
    def valid_partner_name(cls, input_partner_name ):

        
        partners = [name for name in os.listdir('/app/partners/')
                    if os.path.isdir(os.path.join('/app/partners/', name))]


        if input_partner_name in  partners:

            return True
        else:
            return None

###################################################################################################################################
# This function will create and return a spark session
###################################################################################################################################

    @classmethod
    def get_spark_session(cls, appName):
        try:
            spark.stop()
        except:
            None

        spark = SparkSession.builder.master("spark://master:7077").appName(appName).getOrCreate()
        spark.sparkContext.setLogLevel('OFF')

        return spark

###################################################################################################################################
    @classmethod
    def print_failure_message(cls, folder,partner, job, text):


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime


        while len(message) < 39:
            message = message + ' '


        message = message + "--Partner: " + partner

    
        while len(message) < 55:
            message = message + ' '


        message = message + "--Folder: " + folder

    
        while len(message) < 80:
            message = message + ' '
        

        message = message+' --Running: '+ job+ ' Failed!!!' 

        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'

        if "Path does not exist" in text:

            color = 'warning orange'

        else:

            color = 'danger red'

        cls.print_with_style(message, color)
        cls.print_with_style(text, color)




###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_max_feild_size_from_df(cls, file_path,field_name, spark):


        table_df = cls.spark_read(file_path, spark)

        max_length = str(table_df.select(max(length(table_df[field_name]))).collect()[0][0])


        try:
            if int(max_length) >= 1:
               pass
        except:
            max_length = "1"


        return max_length





###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_field_config(cls, file_path, field_name, field_data_type, spark):

        harmonized_size =0 

        field_max_size = int(cls.get_max_feild_size_from_df(file_path,field_name, spark))


        if 'RDBMS Text' in field_data_type:

                field_type = StringType()

                cdm_field_size_list = field_data_type.split('(')

                cdm_field_size = cdm_field_size_list[len(cdm_field_size_list)-1].replace(')','')

                if 'x' in cdm_field_size:
                    
                    if 'RAW' in field_name:


                        cdm_field_size  = cls.get_max_feild_size_from_df(file_path,field_name, spark)


                    else:
                    

                        cdm_field_size = harmonized_filed_sizes_dict.get(field_name,0)
                
                

                if int(cdm_field_size) > int(field_max_size):
                        field_size = cdm_field_size
                else:
                        field_size = field_max_size





                field_config = StructField(field_name, field_type, metadata={"maxlength":int(field_size)})


                    

        if 'RDBMS Date' in field_data_type:

            field_type = DateType()

            field_config =   StructField(field_name, field_type, True)   

        if 'RDBMS Number' in field_data_type:

            field_type = DoubleType() 

            field_config =    StructField(field_name, field_type, True)   


        return field_config


###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_schema_line(cls, file_path, field_name, field_data_type, spark):

        harmonized_size =0 

        field_max_size = int(cls.get_max_feild_size_from_df(file_path,field_name, spark))


        if 'RDBMS Text' in field_data_type:

                schema_type = 'VARCHAR'

                cdm_field_size_list = field_data_type.split('(')

                cdm_field_size = cdm_field_size_list[len(cdm_field_size_list)-1].replace(')','')

                if 'x' in cdm_field_size:
                    
                    if 'RAW' in field_name:


                        cdm_field_size  = cls.get_max_feild_size_from_df(file_path,field_name, spark)

                    else:
                    

                        cdm_field_size = harmonized_filed_sizes_dict.get(field_name,0)
                


                if int(cdm_field_size) > int(field_max_size):
                        field_size = cdm_field_size
                else:
                        field_size = field_max_size



                schema_line =     field_name +' '+ schema_type+"("+str(field_size)+"),\n"


                    

        if 'RDBMS Date' in field_data_type:

            schema_type = 'DATE'

            schema_line =     field_name +' '+ schema_type+",\n"

        if 'RDBMS Number' in field_data_type:

            schema_type = 'FLOAT'

            schema_line =     field_name +' '+ schema_type+",\n"


        return schema_line

        


        


###################################################################################################################################
# This function will adjust the schema fields depend on the size specified by the user or the max size in the field
###################################################################################################################################
    @classmethod
    def get_schemas(cls, file_name, table_path):






        file_path = table_path+'fixed_'+file_name.lower()+".csv"

        schema_mssql = ""
        schema_sf    = []
        schemas      = []
        cdm_fields   = []

        spark = cls.get_spark_session("Get Schema")


        cdm_df = PcornetCDM.get_cdm_df(spark)

        table_cdm_df = cdm_df.filter(cdm_df["TABLE_NAME"]==file_name)

        table_cdm = table_cdm_df.collect()

        fixed_table_df = cls.spark_read(file_path,spark)
  


        for row in table_cdm:


            schema_mssql = schema_mssql + '\n' + cls.get_schema_line(file_path,row["FIELD_NAME"],row["RDBMS_DATA_TYPE"] , spark) 
            schema_sf.append( cls.get_field_config(file_path,row["FIELD_NAME"],row["RDBMS_DATA_TYPE"] , spark) )
            cdm_fields.append(row["FIELD_NAME"])



        

        schema_mssql = schema_mssql +"UPDATED VARCHAR(19),\n"
        schema_mssql = schema_mssql +"SOURCE VARCHAR(20)"
        schema_sf.append(StructField("UPDATED", StringType(), metadata={"maxlength":19}),)
        schema_sf.append(StructField("SOURCE", StringType(), metadata={"maxlength":20}),)

        

        for field in fixed_table_df.columns:

            if field not in cdm_fields and field.upper() not in ['SOURCE','UPDATED']:

                field_size = int(cls.get_max_feild_size_from_df(file_path, field, spark))

                schema_sf.append(StructField(field, StringType(), metadata={"maxlength":field_size}),)
                schema_mssql = schema_mssql +F",\n{field} VARCHAR({field_size})"



        schemas =  [schema_mssql,schema_sf]

        spark.stop()

        return schemas


###################################################################################################################################
# This function will verify the passed value is a float or it will return null
###################################################################################################################################
    @classmethod

    def return_if_float(cls, val):
        try:
            float(val)
            return val
        except:
            #Not a float
            return None
    



###################################################################################################################################
# This function will get all the created mappers python script and return a list of thier names
###################################################################################################################################
    @classmethod
    def print_run_status(cls,job_num,total_jobs_count, job, folder, partner):
        


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime

        while len(message) < 25:
            message = message + ' '

        message = message + 'Job: '+str(job_num)+'/'+str(total_jobs_count)
        while len(message) < 39:
            message = message + ' '

        message = message + "--Partner: " + partner

    
        while len(message) < 55:
            message = message + ' '


        message = message + "--Folder: " + folder

    
        while len(message) < 80:
            message = message + ' '
        

        message = message+' --Running: '+job 

        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'


        cls.print_with_style(message, 'matrix green')

###################################################################################################################################
# This function will print a text with matrix green color
###################################################################################################################################


    @classmethod
    def print_with_style(cls, text, color):

        if color == 'matrix green':

            bold_code = '\033[1m'
            color_code = '\033[92m'
            reset_code = '\033[0m'  # Reset to default style
       
        if color == 'danger red':

            bold_code = '\033[1m'
            color_code = RED='\033[0;31m'
            reset_code = '\033[0m'  # Reset to default style


        if color == 'warning orange':

            bold_code = '\033[1m'
            color_code = '\033[38;5;214m'  # Orange
            reset_code = '\033[0m'  # Reset to default style


        if color == 'blue':
            bold_code = '\033[1m'
            color_code = BLUE='\033[0;34m'
            reset_code = '\033[0m'  # Reset to default style

        if color == 'pomegranate':
            bold_code = '\033[1m'
            color_code =  POMEGRANATE='\033[0;91m'
            reset_code = '\033[0m'  # Reset to default style

        if color == 'dark pink':
            bold_code = '\033[1m'
            color_code =   DARK_PINK='\033[0;35m'
            reset_code = '\033[0m'  # Reset to default style
   
        # Print bold green text
        print(bold_code + color_code + text + reset_code)

###################################################################################################################################
# This function will print a text with matrix green color
###################################################################################################################################


    @classmethod
    def print_fixer_status(cls,current_count,total_count, fix_type, fix_name):
        


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime

        while len(message) < 25:
            message = message + ' '

        message = message + 'Fix: '+str(current_count)+'/'+str(total_count)
        while len(message) < 39:
            message = message + ' '

        message = message + "--Fix type: " + fix_type+ ' '

    
        while len(message) < 80:
            message = message + ' '


        message = message + "--Fix Name: " +  fix_name+ ' '


        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'

        cls.print_with_style(message, 'dark pink')

##################################################################################################################################
# 
###################################################################################################################################

    @staticmethod
    def get_date_from_date_str_with_default_value(  input_date):

            try:
                
                return  str(datetime.strptime(input_date, "%Y-%m-%d").date())
            except:
                return '1900-01-01'        

 
###################################################################################################################################
# T
###################################################################################################################################



   
    @staticmethod
    def get_datetime_from_date_str_with_default_value(  input_date):
        
        try:
            
            return str(datetime.strptime(input_date+ " 00:00:00","%Y-%m-%d %H:%M:%S"))
        except:
            return '1900-01-01 00:00:00'    
        

# ###################################################################################################################################
# # T
# ###################################################################################################################################



   
    @classmethod
    def display_logo(  cls):
        

        img = Image.open('common/ovid_logo.png')
        img.show() 





