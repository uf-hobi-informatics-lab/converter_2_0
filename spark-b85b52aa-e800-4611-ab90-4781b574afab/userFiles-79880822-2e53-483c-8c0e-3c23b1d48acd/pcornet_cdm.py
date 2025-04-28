from openpyxl import load_workbook
import sys



pconrnet_cdm_location = "/app/common/cdm/"
pcornet_cdm_file_name = "pcornet_cdm_v70.csv"
pcornet_valueset_file_name = "2025_01_23_PCORnet_Common_Data_Model_v7dot0_parseable.xlsx"


pcornet_cdm_file_path = pconrnet_cdm_location + pcornet_cdm_file_name
# pcornet_cdm_data = pd.read_csv(pcornet_cdm_file_path)


class PcornetCDM:


    

       


###################################################################################################################################
# This function will return the list of fields from the PCORnet CDM based on the passed table name
###################################################################################################################################

   

    @classmethod
    def get_cdm_valueset_df(cls,field_name):
            


            wb = load_workbook(pconrnet_cdm_location+pcornet_valueset_file_name)
            ws = wb["VALUESETS"]
            header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}


            # if "Field name" in header and "valueset items" in header:
            field_col = header["FIELD_NAME"]
            valueset_col = header["VALUESET_ITEM"]
        

            valueset_list = [
                row[valueset_col - 1]
                for row in ws.iter_rows(min_row=2, values_only=True)
                if row[field_col - 1] == field_name and row[valueset_col - 1] is not None
            ]


            # Close the workbook
            wb.close()

            return  valueset_list


   

    @classmethod
    def get_cdm_df(cls, spark):
            


            cdm_table_df = spark.read.load(pcornet_cdm_file_path,format="csv", sep=",", inferSchema="false", header="true",  quote= '"')
       

            return cdm_table_df












