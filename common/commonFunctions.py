
###################################################################################################################################
# This class will host all the common functions used in the mappings by diffrenet mapping scripts
###################################################################################################################################



from itertools import chain
from Crypto.Hash import SHA
from Crypto.Cipher import XOR
from base64 import b64decode, b64encode
from pyspark.sql.functions import udf, col, count, desc, upper, round,  lit, length, max
from pyspark.sql import SparkSession 
from pyspark import SparkConf, SparkContext
import pyspark
import time
import subprocess 
from pyspark.sql.types import StructType, StructField, StringType, NullType
import sys
from spark_secrets import * 
from settings import *
from pcornet_cdm import PcornetCDM
from pyspark.conf import SparkConf
import pandas as pd
from urllib import parse
from datetime import datetime
import pytz
from dateutil import parser as dp
import pyspark.sql.functions as F
# import importlibcc
import os
import importlib
import openpyxl
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows



POOL_SIZE      = 10
MAX_OVERFLOW   = 5
POOL_RECYCLE   = 3600
BCP_DRIVER     = 'mssql'
  

class CommonFuncitons:


    def __init__(self, partner):




        self.partner = partner # specifying the partner / used for the hashing
        # print(self.partner)
        # sys.exit()
        self.get_time_from_datetime_udf = udf(self.get_time_from_datetime, StringType())
        self.get_date_from_datetime_udf = udf(self.get_date_from_datetime, StringType())
        self.format_date_udf = udf(self.format_date, StringType())
        self.format_time_udf = udf(self.format_time, StringType())
        self.encrypt_id_udf = udf(self.encrypt_id, StringType())
        self.get_current_time_udf = udf(self.get_current_time, StringType())
        self.get_date_from_date_str_with_default_value_udf = udf(self.get_date_from_date_str_with_default_value, StringType())
        self.get_datetime_from_date_str_with_default_value_udf = udf(self.get_datetime_from_date_str_with_default_value, StringType())
        self.copy_if_float_udf= udf(self.copy_if_float, StringType())

        






    


###################################################################################################################################
# This function will return the xor hashed value based on the partner name and input value/id
###################################################################################################################################

   
   

    # @classmethod
    def encrypt_id(cls, id):

            if id is None:

                return None
            else: 


                partner_encryption_value_dict = {

                        'omop_partner_1'   :  'omop_partner_1',
                        'pcornet_partner_1':  'pcornet_partner_1',


                }

                partner_encryption_value = partner_encryption_value_dict.get(cls.partner.upper(), None)

                HASHING_KEY = SHA.new(SEED.encode('utf-8')).digest()
                HASHING_KEY = SHA.new(SEED.encode('utf-8')).digest()

                cipher = XOR.new(HASHING_KEY)
                cipher = XOR.new(HASHING_KEY)

                return b64encode(cipher.encrypt('{}{}'.format(partner_encryption_value, id))).decode()


###################################################################################################################################
# This function will return the current datetime
###################################################################################################################################

   

    @classmethod
    def get_current_time(cls):

       # Set the timezone to Eastern Time (New York)
        eastern_timezone = pytz.timezone("US/Eastern")

        # Get the current time in the Eastern Time Zone
        current_time = datetime.now(eastern_timezone)

        # Define the desired time format
        time_format = "%Y-%m-%d %H:%M:%S"

        # Format and print the current time
        formatted_time = current_time.strftime(time_format)

        return formatted_time


        
    
###################################################################################################################################
# This function will output a pyspark dataframe and combine all the sub files into one file and delete the temprary files
###################################################################################################################################

   

    @classmethod
    def write_pyspark_output_file(cls,payspark_df, output_file_name, output_data_folder_path ):
        pass

        # pcornet_cdm = PcornetCDM()

        # print(pcornet_cdm.get_cdm_table_fields_list('demographic'))
    def write_pyspark_output_file(cls,payspark_df, output_file_name, output_data_folder_path ):

        # pcornet_cdm = PcornetCDM()

        # print(pcornet_cdm.get_cdm_table_fields_list('demographic'))

        date_format_pattern = "yyyy-MM-dd"
        work_directory = output_data_folder_path+"tmp_"+output_file_name

        payspark_df = payspark_df.select([
            F.lit(None).cast('string').alias(i.name)
            if isinstance(i.dataType, NullType)
            else i.name
            for i in payspark_df.schema
            ])
        
        payspark_df.coalesce(1).write.format("csv").option("header", True).option("nullValue",None).option("delimiter", "\t").mode("overwrite").save(work_directory)

        

        command = ["cp", work_directory+"/part-*.csv", output_data_folder_path+output_file_name]
        # Execute the command
        subprocess.run(" ".join(command), shell=True)

        command = ["rm","-r", work_directory]
        # Execute the command
        subprocess.run(" ".join(command), shell=True)

###################################################################################################################################
# This function will 
###################################################################################################################################

           
    def copy_data(cls,input_file_name,output_file_name,data_folder_path,output_data_folder_path ):

        if not os.path.exists(output_data_folder_path):
            os.makedirs(output_data_folder_path)

        command = ["cp", "-p", os.path.join(data_folder_path, input_file_name), os.path.join(output_data_folder_path,output_file_name)]
                # Execute the command
        subprocess.run(" ".join(command), shell=True)

###################################################################################################################################
# This function will 
###################################################################################################################################

    @classmethod

    def apply_fix(cls,fix_type, fix_name, version, input_path, output_path,fixed_table_name,mapped_table_name,input_data_folder,partner_name, src1, src2):
        # print(fix_type)

        if fix_type == 'custom':
            # print('custom')

            partner_fix_path = f"/app/partners/{partner_name.lower()}/custom_fixes/{fix_name}/{version}.py"


        if fix_type == 'common':

            # print('common')
            # cf.print_run_status(0, 0, f'{table}_deduplicator.py', os.path.split(input_folder)[1], partner)

            partner_fix_path = f"/app/common/common_fixes/{fix_name}/{version}.py"

        command = ["python", partner_fix_path, '-f', input_data_folder, '-t', mapped_table_name , '-p', partner_name,'-ftm',fixed_table_name, '-i',input_path,'-o',output_path, '-sr1', src1, '-sr2', src2]

        subprocess.run(" ".join(command), shell=True)




###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   

    @classmethod
    def deduplicate(cls,partner,file_name,table_name, file_path,folder_name):


        # Set up the deduplicated files path
        deduplicated_files_path = file_path.replace('formatter', 'deduplicator')
        duplicate_files_path = os.path.join(deduplicated_files_path, 'duplicate_values')

        # Create the deduplicated files directory if it doesn't exist
        if not os.path.exists(deduplicated_files_path):
            os.makedirs(deduplicated_files_path)

        if not os.path.exists(duplicate_files_path):
            os.makedirs(duplicate_files_path)

        input_file = f"{file_path}formatted_{file_name.lower()}.csv"
        deduplicated_output_file = f"{deduplicated_files_path}deduplicated_{file_name.lower()}.csv"
        duplicates_output_file = f"{duplicate_files_path}/{file_name.lower()}_duplicates.csv"

        # Construct the gawk script based on the file_name
        if file_name == 'ENROLLMENT':
            gawk_script = f'''
            BEGIN {{ FS = OFS = "," }}
            NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
            !seen[$1, $2, $5]++ {{ print > "{deduplicated_output_file}" }}
            seen[$1, $2, $5] > 1 {{ print > "{duplicates_output_file}" }}
            '''
        
        elif file_name == 'DEATH_CAUSE':
            gawk_script = f'''
            BEGIN {{ FS = OFS = "," }}
            NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
            !seen[$1, $2, $3, $4, $5]++ {{ print > "{deduplicated_output_file}" }}
            seen[$1, $2, $3, $4, $5] > 1 {{ print > "{duplicates_output_file}" }}
            '''
        
        else:
            gawk_script = f'''
            BEGIN {{ FS = OFS = "," }}
            NR == 1 {{ header = $0; print header > "{deduplicated_output_file}"; print header > "{duplicates_output_file}"; next }}
            !seen[$1]++ {{ print > "{deduplicated_output_file}" }}
            seen[$1] > 1 {{ print > "{duplicates_output_file}" }}
            '''

        # Construct the full shell command
        command = f"awk '{gawk_script}' {input_file}"
        # Execute the command
        # try:
        subprocess.run(command, shell=True, check=True)
            # print(f"Deduplication completed for {file_name}. Output saved to {output_file}")
        # except subprocess.CalledProcessError as e:
        #     print(f"Error occurred while running gawk: {e}")







###################################################################################################################################
# This function will output the mapping gap for the passed formatted file name
###################################################################################################################################

   

    @classmethod
    def get_mapping_gap(cls,partner,file_name,table_name, file_path,folder_name):


        

        complete_file_path = file_path+'/formatted_'+file_name.lower()+'.csv'
        partner_dictionaries_path = "partners."+partner+".dictionaries"
        partner_dictionaries = importlib.import_module(partner_dictionaries_path)

        mapping_dicionaries_list = []

        for key, value in partner_dictionaries.__dict__.items():
            spark = SparkSession.builder.master("spark://master:7077").appName("MAPPING GAP").getOrCreate()
            # print(key )
            # print(value)

            if  ((file_name.lower() in key and 'death_cause' not in key) or  (file_name.lower() in key and file_name.lower()=='death_cause') ) and "_dict" in key:
                
                # get the field name


                prefix = '@' #logic added to only remove the first word from the dictinary name

                temp_key = prefix+key


                field_name = temp_key.replace(prefix+file_name.lower()+'_',"").replace('_dict',"").upper()

                try:

                    aggregated_values = cls.get_aggregate_value_from_table(complete_file_path,field_name,spark)


                    dictionary_dataframe = spark.createDataFrame(value.items(), ["dict_key", "dict_value"])


                    if 'UNIT' in field_name:
                        joined_df = aggregated_values.join(dictionary_dataframe, dictionary_dataframe['dict_key']==aggregated_values[field_name], how ="left").orderBy(desc("count"))
                    else:
                        joined_df = aggregated_values.join(dictionary_dataframe, upper(dictionary_dataframe['dict_key'])==upper(aggregated_values[field_name]), how ="left").orderBy(desc("count"))

                    cls.diagnos_dictionary(dictionary_dataframe, file_path, key, file_name.lower())

                    cls.daignos_mapping(joined_df, file_path, key, field_name, dictionary_dataframe, file_name.lower())

                    mapping_dicionaries_list.append(key)

                    spark.stop()

                except Exception as e:
                    spark.stop
                    cls.print_failure_message(
                        folder = folder_name,
                        partner = partner.lower(),
                        job = 'mapping gap',
                        text = str(e)
                                            )


###################################################################################################################################
# This function will diagnos a given dictionary data frame
########################################################################################################

    @classmethod
    def diagnos_dictionary(cls,dictionary_dataframe, file_path, file_name, table_name):

        # getting values that mapped to null or some flavor of null like 'OT, 'NI', and "UN
        filtered_df = dictionary_dataframe.filter((col("dict_value").isin("NI", "OT", "UN")) | (col("dict_value").isNull()) | (col("dict_value") == ""))

        out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/mapping_to_flavor_of_null/"
        output_file_name = file_name+"_flavor_of_null.csv"


        cls.write_pyspark_output_file(cls,filtered_df, output_file_name, out_file_path )
           


###################################################################################################################################
# This function will diagnos a given joind dictionary dataframe
########################################################################################################

    @classmethod
    def daignos_mapping(cls,joined_df, file_path, file_name, field_name, dictionary_dataframe, table_name):


        # try:
            # print("*************************** daignos_mapping 1 *************************************")
            # output all mapping
            all_values_df = joined_df.select(joined_df[field_name],joined_df['dict_value'],joined_df['count'],joined_df['percent'])
            # print("*************************** daignos_mapping 2 *************************************")
            all_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/all_mapping/"
            # print("*************************** daignos_mapping 3 *************************************")
            all_mapping_output_file_name = file_name+"_all_mapping.csv"

            cls.write_pyspark_output_file(cls,all_values_df, all_mapping_output_file_name, all_mapping_out_file_path )

            # print("*************************** daignos_mapping 4 *************************************")

            
            # output missing mapping
            missing_values_df = all_values_df.filter(col("dict_value").isNull() )

            missing_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/missing_mapping/"
            missing_mapping_output_file_name = file_name+"_missing_mapping.csv"
            cls.write_pyspark_output_file(cls,missing_values_df, missing_mapping_output_file_name, missing_mapping_out_file_path )
            # print("*************************** daignos_mapping 5 *************************************")

            # output suggested mapping

            
            to_be_added_field = missing_values_df.select(

                    missing_values_df[field_name].alias('dict_key'),
                    lit('').alias('dict_value'),


            )
            # print("*************************** daignos_mapping 6 *************************************")

            unioned_df= dictionary_dataframe.union(to_be_added_field)
            # unioned_df.show()
            suggested_mapping_out_file_path = file_path.replace('formatter','mapping_gap')+"/"+table_name+"/suggested_mapping/"
            # print("*************************** daignos_mapping 7 *************************************")

            out_file_name = suggested_mapping_out_file_path+"suggested_"+file_name+".py"

            result = os.makedirs(os.path.dirname(out_file_name), exist_ok=True)

            # print("*************************** daignos_mapping 8 *************************************")
            # unioned_df.show()
            formatted_text = unioned_df.rdd \
                .map(lambda row: f'       {cls.process_key(row["dict_key"],row["dict_value"],file_name, 20)}  :  "{row["dict_value"]}",') 
            
            # print("*************************** daignos_mapping 8 -2 *************************************")

            formatted_text= formatted_text.collect()
            
            # print("*************************** daignos_mapping 8 -2 *************************************")

            # Join the formatted strings with newline characters
            formatted_text = "\n".join(formatted_text)
            # print("*************************** daignos_mapping 9 *************************************")

            # Write the formatted text to a text file
        
            first_line = "\n\n"+file_name+ " = {\n\n"
            last_line = "   \n\n}"
            # print("*************************** daignos_mapping 10 *************************************")

            with open(out_file_name, "w+") as file:
                file.write(first_line)
                file.write(formatted_text)
                file.write(last_line)


        # except Exception as e:

            

        #     cls.print_with_style(str(e), 'danger red')
            
            
    
###################################################################################################################################
# 
########################################################################################################

    @classmethod
    def process_row(cls,row, file):


            print(row)


            key = cls.append_spaces(row.dict_key, 20),

            text = row.dict_value + '\n'

            # text = row['dict_value']

            file.write(text)


                
               

###################################################################################################################################
# 
########################################################################################################

    @classmethod
    def process_key(cls, key, value,file_name, size):

            key = str(key)

            if 'unit' not in file_name and 'UNIT' not in file_name:

                key = key.upper()

            key = '"'+key+'"'


            if value == "" or value == None:

                key = '# '+key

            while len(key) < size:

                key = key + ' '

            return key





###################################################################################################################################
# This function return and aggreated dataframe for a given table
###################################################################################################################################


    @classmethod
    def get_aggregate_value_from_table(cls,file_path,field_name,spark):

        df = cls.spark_read(file_path, spark)
        row_count= df.count()

        aggregated_df = df.groupBy(field_name).agg(count("*").alias("count"))
        
        if row_count != 0 :

                    aggregated_df_with_percent= aggregated_df.select(

                    aggregated_df[field_name],
                    aggregated_df['count'],
                    
                    round((aggregated_df['count']/row_count)*100,2).alias('percent')


                             )   
            
        else :

                    aggregated_df_with_percent= aggregated_df.select(

                    aggregated_df[field_name],
                    aggregated_df['count'],
                    
                    lit("0").alias('percent')
                    )


        return aggregated_df_with_percent




###################################################################################################################################
# This function will upload a pyspark df to mssql table in a specified server/dabase
###################################################################################################################################

   

    @classmethod
    def db_upload(cls,db, db_server,schema,file_name,table_name, file_path,folder_name):

        date_format_pattern = "yyyy-MM-dd"

        spark = SparkSession.builder.master("spark://master:7077").appName("MSSQL Upload").getOrCreate()

        try:
            conf = SparkConf()
            conf.set("spark.driver.extraClassPath", "mssql-jdbc-driver.jar")
            file_df = spark.read.option("inferSchema", "false").load(file_path+file_name+'*',format="csv", sep="\t", inferSchema="false", header="true",  quote= '"')
            
            # string_columns = [f'CAST({col} AS STRING)' for file_df in file_df.columns]
            string_columns = [f'CAST(`{col}` AS STRING) AS `{col}`' for col in file_df.columns]
            result_df = file_df.selectExpr(*string_columns)

            
            jdbc_url = f"jdbc:sqlserver://{db_server};databaseName={db}"
            
            if schema != None:
                properties = {
                    "user": DB_USER,
                    "password": DB_PASS,
                    "driver": "com.microsoft.sqlserver.jdbc.SQLServerDriver",
                    "trustServerCertificate": "true",
                    "createTableColumnTypes": schema}
            else:
                properties = {
                    "user": DB_USER,
                    "password": DB_PASS,
                    "driver": "com.microsoft.sqlserver.jdbc.SQLServerDriver",
                    "trustServerCertificate": "true"}


            table_name = f"PYSPARK_{folder_name}_{table_name}"
            
            result_df.write.jdbc(url=jdbc_url, table=table_name, mode="overwrite", properties=properties)

            spark.stop()
        except Exception as e:

            spark.stop()
            

            cls.print_with_style(str(e), 'danger red')



###################################################################################################################################
# This function will return a list of folder in a given path
###################################################################################################################################
    @classmethod
    def get_folders_list(cls, path):

       
        
        folders_names = os.listdir(path)

        folders_names_list = [file for file in folders_names if file]

        return folders_names_list


###################################################################################################################################
# This function generate a mapping report as a form of an excel file
###################################################################################################################################


    @classmethod
    def generate_mapping_report(cls,mapping_gap_output_folder_path, folder, partner):

        spark = cls.get_spark_session('generate mapping report')

        empty_columns = pd.DataFrame(columns=[""] * 2)
        # Path to the Excel file
        excel_file_path = mapping_gap_output_folder_path + partner+'_'+folder+"_mapping_report.xlsx"

        # Initialize a workbook object
        wb = openpyxl.Workbook()



        tables_list = cls.get_folders_list(mapping_gap_output_folder_path)

        for table in tables_list :

           


            if 'mapping_report' not in table:

                ws = wb.create_sheet(title=table)

                combined_pd = empty_columns


                table_all_mappings_path = mapping_gap_output_folder_path +"/"+table+"/all_mapping/"

                table_all_mapping_files_list = cls.get_folders_list(table_all_mappings_path)

                for mapping_file in table_all_mapping_files_list:

                    mapping_df =  cls.spark_read(table_all_mappings_path+mapping_file, spark)

                    mapping_df = mapping_df.limit(1000000)


                    pandas_df = mapping_df.toPandas()
                    combined_pd = pd.concat([combined_pd, empty_columns, pandas_df], axis=1)


                    for col_idx, col_name in enumerate(combined_pd.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)


                    for row_idx, row in enumerate(combined_pd.values, 2):  # Start from row 2 (after header row)
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)

                            # Calculate the width of the cell based on the length of the value
                            if isinstance(value, str):
                                width = len(value)
                            else:
                                width = len(str(value))

                            # Update the column width if necessary
                            if width > ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width:
                                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


        if 'Sheet' in wb.sheetnames:
            # Remove the sheet
            wb.remove(wb['Sheet'])

        wb.save(excel_file_path)




        spark.stop()



###################################################################################################################################
# This function will return a pysark dataframe for a givin file path
###################################################################################################################################


    @classmethod
    def spark_read(cls,file_path, spark):

        return spark.read.option("inferSchema", "false").load(file_path,format="csv",   sep="\t", inferSchema="false", header="true",  quote= '"')



###################################################################################################################################
# This function will append any additional fields that are not in the Pcornet CDM to the mapper output
###################################################################################################################################

   

    @classmethod
    def append_additional_fields(cls,mapped_df, file_name, deduplicated_data_folder_path,join_field, spark ):


        # read the deduplicated file
        deduplicated_file = spark.read.load(deduplicated_data_folder_path+file_name,format="csv", sep="\t", inferSchema="true", header="true", quote= '"')
        
        #get the headers from the deduplicated and the mapped files for comparison
        deduplicated_file_columns = deduplicated_file.columns
        mapped_df_columns = mapped_df.columns


        # Identify common headers
        common_headers = set(mapped_df_columns).intersection(deduplicated_file_columns)
        common_headers_without_the_join_field = [item for item in common_headers if item != join_field]

        # Drop common headers from the second DataFrame
        additional_fields_data = deduplicated_file.drop(*common_headers_without_the_join_field)
        
        #Renaming the join field to avoid conflect during the join
        join_field_temp_name = join_field+"_new"

        additional_fields_data = additional_fields_data.withColumnRenamed(join_field, join_field_temp_name)
        
        # Join DataFrames based on the "Name" column
        appended_payspark_data_frame = mapped_df.join(additional_fields_data, mapped_df["JOIN_FIELD"]==additional_fields_data[join_field_temp_name])
       
       #Dropping the temporary join fields

        columns_to_drop = ["JOIN_FIELD",join_field_temp_name ]
        appended_payspark_data_frame = appended_payspark_data_frame.drop(*columns_to_drop)

        return appended_payspark_data_frame
        


###################################################################################################################################
# This function will convert all known time formats to the '%H:%M'format
###################################################################################################################################

    @staticmethod
    def format_time( val):
        """ Convenience method to try all known formats """

        try:
            parsed_time = dp.parse(val)
        except Exception as exc:
           
            return ""
        return parsed_time.strftime('%H:%M')


###################################################################################################################################
# This will ensure the value is a float or it will return None
###################################################################################################################################

    @staticmethod
    def copy_if_float( val):
        """ Convenience method to try all known formats """
        if not val:
            return None
        try:
            return float(val)
        except:

            return None



###################################################################################################################################
# This function will convert all known date formats to the '%Y-%m-%d format
###################################################################################################################################

   
    @staticmethod
    def format_date( val):
        """ Convenience method to try all known formats """
        print(val)
        try:
            parsed_date = dp.parse(val)
        except:

            print(val)
           
            return None

        return parsed_date.strftime('%Y-%m-%d')




###################################################################################################################################
# This function will return the time portion from a datetime value
###################################################################################################################################

   
    @staticmethod
    def get_time_from_datetime(  dateTime):
        
            try:
                
                return str(dateTime.time().strftime('%H:%M:%S'))[0:5]
            except:
                return None

###################################################################################################################################
# This function will return the date portion from a datetime value
###################################################################################################################################

   
    @staticmethod
    def get_date_from_datetime(  dateTime):
        
        try:
            
            return str(dateTime.strftime("%Y-%m-%d"))
        except:
            return None

        
  
###################################################################################################################################
# This function will validate if a giving partenr name is valid or not
###################################################################################################################################

   

    @classmethod
    def valid_partner_name(cls, input_partner_name ):

        

        if input_partner_name in  partners_list:

            return True
        else:
            return None

###################################################################################################################################
# This function will create and return a spark session
###################################################################################################################################

    @classmethod
    def get_spark_session(cls, appName):
        try:
            spark.stop()
        except:
            None

        spark = SparkSession.builder.master("spark://master:7077").appName(appName).getOrCreate()
        spark.sparkContext.setLogLevel('OFF')

        return spark

###################################################################################################################################
    @classmethod
    def print_failure_message(cls, folder,partner, job, text):


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime


        while len(message) < 39:
            message = message + ' '


        message = message + "--Partner: " + partner

    
        while len(message) < 55:
            message = message + ' '


        message = message + "--Folder: " + folder

    
        while len(message) < 80:
            message = message + ' '
        

        message = message+' --Running: '+ job+ ' Failed!!!' 

        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'

        if "Path does not exist" in text:

            color = 'warning orange'

        else:

            color = 'danger red'

        cls.print_with_style(message, color)
        cls.print_with_style(text, color)




###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_max_feild_size_from_df(cls, file_path,field_name, spark):

        # print("inside get_max_feild_size_from_df")

        table_df = cls.spark_read(file_path, spark)

        # table_df.show()
        # print("field_name")
        # print(field_name)
        max_length = str(table_df.select(max(length(table_df[field_name]))).collect()[0][0])

        # print("max_length")
        # print(max_length)

        try:
            if int(max_length) >= 1:
               pass
        except:
            max_length = "1"


        return max_length




###################################################################################################################################
# This 
###################################################################################################################################
    @classmethod
    def get_schema_line(cls, file_path, field_name, field_data_type, spark):
         

        # print ("get_schema_line  for : #######################################################")
        # print (field_name)
        harmonized_size =0 

        field_max_size = int(cls.get_max_feild_size_from_df(file_path,field_name, spark))
        # print("field_max_size")
        # print(field_max_size)

        if 'RDBMS Text' in field_data_type:

                schema_type = 'VARCHAR'

                cdm_field_size_list = field_data_type.split('(')

                cdm_field_size = cdm_field_size_list[len(cdm_field_size_list)-1].replace(')','')

                if 'x' in cdm_field_size:
                    
                    if 'RAW' in field_name:

                        # print("outside  RAW get_max_feild_size_from_df")

                        cdm_field_size  = cls.get_max_feild_size_from_df(file_path,field_name, spark)

                    else:
                        # print("outside  ELSE get_max_feild_size_from_df")

                        cdm_field_size = harmonized_filed_sizes_dict.get(field_name,0)
                


                # print('cdm_field_size')
                # print(cdm_field_size)
                # print('field_max_size')
                # print(field_max_size)


                if int(cdm_field_size) > int(field_max_size):
                        field_size = cdm_field_size
                else:
                        field_size = field_max_size


                        # cdm_field_size  = int(harmonized_filed_sizes_dict.get(field_name))
                                              
                        # if int(harmonized_filed_sizes_dict.get(field_name)) <   int(cls.get_max_feild_size_from_df(file_path,field_name, spark)):
                        
                        #                     cdm_field_size  =  cls.get_max_feild_size_from_df(file_path,field_name, spark)

                # print('field_name')
                # print(field_name)

                # print('harmonized_size')
                # print(harmonized_size)
                
                # print('field_max_size')
                # print(field_max_size)

                # print('cdm_field_size')
                # print(cdm_field_size)


                schema_line =     field_name +' '+ schema_type+"("+str(field_size)+"),\n"


                    

        if 'RDBMS Date' in field_data_type:

            schema_type = 'DATE'

            schema_line =     field_name +' '+ schema_type+",\n"

        if 'RDBMS Number' in field_data_type:

            schema_type = 'FLOAT'

            schema_line =     field_name +' '+ schema_type+",\n"


        return schema_line

        


        

###################################################################################################################################
# This function will adjust the schema fields depend on the size specified by the user or the max size in the field
###################################################################################################################################
    # @classmethod
    # def remove_last_comma(cls,s):
    # # Find the last occurrence of the comma
    #     last_comma_index = s.rfind(',')
    #     # If there is no comma in the string, return the string as is
    #     if last_comma_index == -1:
    #         return s
    #     # Remove the last comma by slicing the string
    #     return s[:last_comma_index] + s[last_comma_index + 1:]

###################################################################################################################################
# This function will adjust the schema fields depend on the size specified by the user or the max size in the field
###################################################################################################################################
    @classmethod
    def get_schema(cls, file_name, table_path):

        # print("*********** get_schema ")
        # print(file_name)

        file_path = table_path+'fixed_'+file_name.lower()+".csv"

        schema = ""

        spark = cls.get_spark_session("Get Schema")


        cdm_df = PcornetCDM.get_cdm_df(spark)

        table_cdm_df = cdm_df.filter(cdm_df["TABLE_NAME"]==file_name)

        table_cdm = table_cdm_df.collect()

        for row in table_cdm:
            # print(row["FIELD_NAME"])
            # print(row["RDBMS_DATA_TYPE"])

            schema = schema + '\n' + cls.get_schema_line(file_path,row["FIELD_NAME"],row["RDBMS_DATA_TYPE"] , spark) 


        
        schema = schema +"UPDATED VARCHAR(19),\n"
        schema = schema +"SOURCE VARCHAR(20)"

        # schema = cls.remove_last_comma(schema)



        # cdm_df = 

        # schema_list = schema.split(",")


        # for line in schema_list:
        #     line.replace ''
        # print(schema)

        spark.stop()
        return schema





###################################################################################################################################
# This function will get all the created mappers python script and return a list of thier names
###################################################################################################################################
    @classmethod
    def print_run_status(cls,job_num,total_jobs_count, job, folder, partner):
        


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime

        while len(message) < 25:
            message = message + ' '

        message = message + 'Job: '+str(job_num)+'/'+str(total_jobs_count)
        while len(message) < 39:
            message = message + ' '

        message = message + "--Partner: " + partner

    
        while len(message) < 55:
            message = message + ' '


        message = message + "--Folder: " + folder

    
        while len(message) < 80:
            message = message + ' '
        

        message = message+' --Running: '+job 

        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'


        cls.print_with_style(message, 'matrix green')

###################################################################################################################################
# This function will print a text with matrix green color
###################################################################################################################################


    @classmethod
    def print_with_style(cls, text, color):

        if color == 'matrix green':

            bold_code = '\033[1m'
            color_code = '\033[92m'
            reset_code = '\033[0m'  # Reset to default style
       
        if color == 'danger red':

            bold_code = '\033[1m'
            color_code = RED='\033[0;31m'
            reset_code = '\033[0m'  # Reset to default style


        if color == 'warning orange':

            bold_code = '\033[1m'
            color_code = '\033[38;5;214m'  # Orange
            reset_code = '\033[0m'  # Reset to default style


        if color == 'blue':
            bold_code = '\033[1m'
            color_code = BLUE='\033[0;34m'
            reset_code = '\033[0m'  # Reset to default style

        if color == 'pomegranate':
            bold_code = '\033[1m'
            color_code =  POMEGRANATE='\033[0;91m'
            reset_code = '\033[0m'  # Reset to default style

        if color == 'dark pink':
            bold_code = '\033[1m'
            color_code =   DARK_PINK='\033[0;35m'
            reset_code = '\033[0m'  # Reset to default style
   
        # Print bold green text
        print(bold_code + color_code + text + reset_code)

###################################################################################################################################
# This function will print a text with matrix green color
###################################################################################################################################


    @classmethod
    def print_fixer_status(cls,current_count,total_count, fix_type, fix_name):
        


        current_utc_datetime = datetime.utcnow()
        new_york_tz = pytz.timezone('America/New_York')

        current_ny_datetime = current_utc_datetime.replace(tzinfo=pytz.utc).astimezone(new_york_tz)

        formatted_ny_datetime = current_ny_datetime.strftime("%Y-%m-%d %H:%M:%S %Z")

        message = formatted_ny_datetime

        while len(message) < 25:
            message = message + ' '

        message = message + 'Fix: '+str(current_count)+'/'+str(total_count)
        while len(message) < 39:
            message = message + ' '

        message = message + "--Fix type: " + fix_type+ ' '

    
        while len(message) < 80:
            message = message + ' '


        message = message + "--Fix Name: " +  fix_name+ ' '


        while len(message) < 125:
            message = message + ' '

        while len(message) < 145:
            message = message + '.'

        cls.print_with_style(message, 'dark pink')

##################################################################################################################################
# 
###################################################################################################################################

    @staticmethod
    def get_date_from_date_str_with_default_value(  input_date):

            try:
                
                return  str(datetime.strptime(input_date, "%Y-%m-%d").date())
            except:
                return '1900-01-01'        

 
###################################################################################################################################
# T
###################################################################################################################################



   
    @staticmethod
    def get_datetime_from_date_str_with_default_value(  input_date):
        
        try:
            
            return str(datetime.strptime(input_date+ " 00:00:00","%Y-%m-%d %H:%M:%S"))
        except:
            return '1900-01-01 00:00:00'    